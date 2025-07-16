import time
from openpyxl import load_workbook, Workbook
from description_generator import get_description
from sqlalchemy import create_engine, MetaData, Table, update, select

SQLALCHEMY_DATABASE_URI = ''
engine = create_engine(SQLALCHEMY_DATABASE_URI, echo=True)
metadata_obj = MetaData()
metadata_obj.reflect(bind=engine)
content = Table("content", metadata_obj, autoload_with=engine)

def import_from_file(filename, sheet):
	workbook = load_workbook(filename=filename)
	sheet = workbook[sheet]
	for row in sheet.iter_rows(min_col=1, max_col=12, min_row=2, values_only=True):
		with engine.connect() as conn:
			update_stmt = update(content).where(content.c.id == row[0]).values(description=row[11])
			conn.execute(update_stmt)

def import_from_imdb():
	with engine.connect() as conn:
		select_stmt = select(content).where(content.c.description == '')
		result = conn.execute(select_stmt)
		for row in result:
			description = get_description(row['title'])
			update_stmt = update(content).where(content.c.id == row['id']).values(description = description)
			conn.execute(update_stmt)
			time.sleep(5)

#import_from_file('IMDB AUDIOVAULT DATA.xlsx', 'tv shows')
#import_from_file('IMDB AUDIOVAULT DATA.xlsx', 'movies')
import_from_imdb()
